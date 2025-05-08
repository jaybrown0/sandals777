import os
import re
import json
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import shutil

# Constants
today_date = pd.Timestamp.today().strftime("%Y-%m-%d")
excel_file_path = r"C:\Git\data.xlsx"

headers = {
    "User-Agent": "Mozilla/5.0"
}

# Scrape Sandals and Beaches Deals
sources = {
    "Sandals": "https://www.sandals.com/specials/suite-deals/",
    "Beaches": "https://www.beaches.com/specials/suite-deals/"
}

all_promotions = set()

# Scrape both sites
for brand, url in sources.items():
    print(f"\n🔍 Fetching: {brand} Suite Deals...")
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print(f"❌ Failed to fetch {brand}. Status: {response.status_code}")
        continue

    soup = BeautifulSoup(response.text, "html.parser")
    script_tags = soup.find_all("script")

    for tag in script_tags:
        content = tag.string
        if content and "promotionTitle" in content:
            matches = re.findall(r'push\(\[\d+,\s*"(.+?)"\]\)', content, re.DOTALL)
            for blob in matches:
                decoded = bytes(blob, "utf-8").decode("unicode_escape")
                json_objects = re.findall(r'\{[^{}]*?"promotionTitle"\s*:\s*"7-7-7[^{}]*?\}', decoded)
                for js in json_objects:
                    try:
                        obj = json.loads(js)
                        title = obj.get("promotionTitle", "")
                        rst = obj.get("rstCode", "N/A")
                        room = obj.get("roomCategory", "N/A")
                        all_promotions.add((title, rst, room))
                    except Exception:
                        continue

# Convert to DataFrame for display only
df = pd.DataFrame(list(all_promotions), columns=["promotionTitle", "rstCode", "roomCategory"])
df_current_promotions = df.sort_values(by=["rstCode"]).reset_index(drop=True)

print("\n✅ Extracted Promotions:\n")
print(df_current_promotions)

# === Excel Handling ===

# Load or create workbook
if os.path.exists(excel_file_path):
    wb = load_workbook(excel_file_path)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "7-7-7 Promotions"
    sheet.append(["Start Date"])  # header row

# Get current header row (excluding "Start Date")
headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]

# Get current start dates
start_dates = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]

# Find or create today's row
if today_date in start_dates:
    row_index = start_dates.index(today_date) + 2
else:
    row_index = sheet.max_row + 1
    sheet.cell(row=row_index, column=1, value=today_date)

# Write promotions
for promotion_title, rst_code, room_category in df.itertuples(index=False):
    # Add new header if needed
    if rst_code not in headers:
        next_col = len(headers) + 2
        sheet.cell(row=1, column=next_col, value=rst_code)
        headers.append(rst_code)

    col_index = headers.index(rst_code) + 2
    existing_value = sheet.cell(row=row_index, column=col_index).value

    if existing_value:
        categories = set(existing_value.split(", "))
        categories.add(room_category)
        sheet.cell(row=row_index, column=col_index, value=", ".join(sorted(categories)))
    else:
        sheet.cell(row=row_index, column=col_index, value=room_category)

# Save
wb.save(excel_file_path)
wb.close()

print(f"\n✅ Data updated in: {excel_file_path}")
